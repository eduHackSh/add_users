import openpyxl
import string





#Variable for iterate the letters, A:B:C:...
abc = string.ascii_uppercase
#Variable for contain the row data
row = []
#Variable for contain a list of rows in dicts
lst = []

def Open_Book():
    book = openpyxl.load_workbook(input("Introduce el nombre del archivo excel:\n"))
    sheet = book.active
    return sheet

#Variable for contain the max row num
def maxRow(sheet):
    return sheet.max_row


    #Iterate the sheet row per row, x is the number of row
def Obtain_Users(sheet):
    for x in range(2, sheet.max_row+1):
        #Iterate the character to match it with the num of row(x) 
        for char in range(0, 6):
            userData = sheet[(abc[char] + str(x))].value
            row.append(userData)
        #Insert the data of the user in a array using a dict    
        lst.append({'nombre': row[0], 'apellido': row[1], 'email': row[2], 'cargo': row[3],
                    'user': row[4], 'password': row[5]}) 

        row.clear()
    return lst



